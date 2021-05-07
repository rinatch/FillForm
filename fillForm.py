import openpyxl
from openpyxl import Workbook
from pathlib import Path
import os
from selenium import webdriver
import time
from selenium.webdriver.support.ui import Select,WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys


STUDENT_CLASS = 'ח'
MALE_GENDER_ARABIC = 'تلميذ'
FEMALE_GENDER_ARABIC = 'تلميذة'
MALE_GENDER_HEBREW = 'תלמיד'
FEMALE_GENDER_HEBREW = 'תלמידה'
web = webdriver.Chrome(os.getcwd()+"/chromedriver.exe")
web.get('https://pm.cyber.org.il/students/new_student_application_nm')
time.sleep(2)
wb = Workbook()
ws = wb.active
ws.title = "Invalid Student Data"
wb.save(filename='invalid_student_data.xlsx')



def iter_rows(ws, n):  # produce the list of items in the particular row
    for row in ws.iter_rows(n):
        yield [cell.value for cell in row]

def read_excel(file_name):
    # Setting the path to the xlsx file:
    path = os.getcwd()
    xlsx_file = Path(path, file_name + '.xlsx')
    try:
        wb_obj = openpyxl.load_workbook(xlsx_file)
    except ValueError:
        print("Error Loading File, not in current directory or name file not correct :(")
    wsheet = wb_obj.active
    return wsheet

invalid_file_idx = 1
def write_excel(row, idx):
     global invalid_file_idx
     new_excel = openpyxl.load_workbook('invalid_student_data.xlsx')
     wnew_excel_sheet = new_excel.active
     for elem,col in zip(row,range(1,12)):
         wnew_excel_sheet.cell(row=invalid_file_idx, column=col).value = elem.value
     invalid_file_idx += 1
     new_excel.save('invalid_student_data.xlsx')

def clear_fields():
    form_first_name = web.find_element_by_xpath('//*[@id="student_first_name"]')
    form_last_name = web.find_element_by_xpath('//*[@id="student_last_name"]')
    form_id = web.find_element_by_xpath('//*[@id="student_p_id"]')
    form_gender = web.find_element_by_xpath('//*[@id="student_gender"]')
    form_phone_number = web.find_element_by_xpath('//*[@id="student_mobile1"]')
    form_email = web.find_element_by_xpath('//*[@id="student_primary_email"]')
    form_city = web.find_element_by_xpath('//*[@id="s2_wrapper"]/span/span[1]/span')
    form_class = web.find_element_by_xpath('//*[@id="student_school_class"]')
    form_parent_name = web.find_element_by_xpath('//*[@id="student_father_name"]')
    form_parent_number = web.find_element_by_xpath('//*[@id="student_father_phone"]')
    form_parent_email = web.find_element_by_xpath('//*[@id="student_father_email"]')
    form_submit = web.find_element_by_xpath('//*[@id="update_student_form"]/div[2]/div[4]/input')

    form_first_name.clear()
    form_last_name.clear()
    form_id.clear()
    form_phone_number.clear()
    form_email.clear()
    form_parent_name.clear()
    form_parent_number.clear()
    form_parent_email.clear()

def check_data_validity(field_type, data):
    if field_type == "id":
        if len(str(data)) != 9:
            return False
    if field_type == "phone number":
        if len(str(data)) == 9:
            return '0' + data
        if str(data).find('972') < 2:
            str(data).replace('972','0') #Fix a place where 972 also in the middle


def fill_examinee_data(sheet):

    i = 2 #in invalid file, i should begin from 1
    # when running the invalid file after fixing it, first param should be 1
    for row in sheet.iter_rows(2, sheet.max_row):
        form_first_name = web.find_element_by_xpath('//*[@id="student_first_name"]')
        form_last_name = web.find_element_by_xpath('//*[@id="student_last_name"]')
        form_id = web.find_element_by_xpath('//*[@id="student_p_id"]')
        form_gender = web.find_element_by_xpath('//*[@id="student_gender"]')
        form_phone_number = web.find_element_by_xpath('//*[@id="student_mobile1"]')
        form_email = web.find_element_by_xpath('//*[@id="student_primary_email"]')
        form_city = web.find_element_by_xpath('//*[@id="s2_wrapper"]/span/span[1]/span')
        form_class = web.find_element_by_xpath('//*[@id="student_school_class"]')
        form_parent_name = web.find_element_by_xpath('//*[@id="student_father_name"]')
        form_parent_number = web.find_element_by_xpath('//*[@id="student_father_phone"]')
        form_parent_email = web.find_element_by_xpath('//*[@id="student_father_email"]')
        form_submit = web.find_element_by_xpath('//*[@id="update_student_form"]/div[2]/div[4]/input')

        select_gender = Select(form_gender)
        select_class = Select(form_class)
        print(i, ") Inserting", row[0].value, row[1].value, row[2].value, "=> "),
        form_first_name.send_keys(row[0].value)
        form_last_name.send_keys(row[1].value)
        form_id.send_keys(row[2].value)
        if row[3].value == MALE_GENDER_ARABIC or row[3].value == MALE_GENDER_HEBREW:
            select_gender.select_by_visible_text('זכר')
        elif row[3].value == FEMALE_GENDER_ARABIC or row[3].value == FEMALE_GENDER_HEBREW:
            select_gender.select_by_visible_text('נקבה')
        else:
            select_gender.select_by_visible_text(row[3].value)
        form_phone_number.send_keys('0'+str(row[4].value))
        form_email.send_keys(row[5].value)
        form_city.click()
        search = web.find_element_by_xpath('/html/body/span/span/span[1]/input')

        search.send_keys(u"דאלית")
        time.sleep(2)
        elems = web.find_element_by_xpath('//*[@id="select2-student_main_city_cat_id-results"]')
        for element in elems.find_elements_by_class_name("select2-results__option"):
            if element.text == "דאלית אל-כרמל":
                element.click()
        time.sleep(2)
        select_school = Select(web.find_element_by_xpath('//*[@id="student_partner_id"]'))
        option = select_school.options
        for index in range(len(option)-1):
            if row[7].value in option[index].text:
                select_school.select_by_visible_text(option[index].text)
        select_class.select_by_visible_text(STUDENT_CLASS)
        if row[8].value:
            form_parent_number.send_keys('0'+str(row[8].value))
        if row[9].value:
            form_parent_email.send_keys(row[9].value)
        form_parent_name.send_keys(row[10].value)
        form_submit.click()
        time.sleep(3)
        try:
            student_submitted_error = web.find_element_by_xpath('//*[@id="errorExplanation"]/h3')
            print(student_submitted_error.text)
            if(student_submitted_error.text == '3 בעיות מונעות את השמירה'):
                clear_fields()
                time.sleep(2)
                i += 1
                continue
            elif student_submitted_error.text == 'בעיה אחת מונעת את השמירה' or student_submitted_error.text == 'שתי בעיות מונעת את השמירה' or student_submitted_error.text == '2 בעיות מונעת את השמירה':
                clear_fields()
                time.sleep(2)
                write_excel(row, i)
                i += 1
                continue
        except:
            pass

        try:
            get_confirmation_div_text = web.find_element_by_xpath('/html/body/div[1]/section/div/div/div/h1')
        except:
            print("Worng input!")
            #refresh
            web.refresh()
            #insert this row to second excel
            write_excel(row, i)
            i += 1
            continue
        if get_confirmation_div_text.text == "טופס הרשמה נקלט בהצלחה":
            print("Successfully Received")
        else:
            print("Unable To Insert Examinee")
        try:
            re_submit = web.find_element_by_xpath('/html/body/div[1]/section/div/div/div/div/div/h2/a')
            re_submit.click()
        except:
            print("Worng input 2!")
            # refresh
            web.refresh()
            # insert this row to second excel
            write_excel(row, i)
            i += 1
            continue
        i += 1

data_sheet = read_excel('examiners_info')
fill_examinee_data(data_sheet)

